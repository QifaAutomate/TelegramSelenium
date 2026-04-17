from pathlib import Path
from openpyxl import Workbook, load_workbook


def read_xlsx(path: Path) -> list[tuple]:
    """Читает Excel и возвращает список строк (без заголовка)."""
    wb = load_workbook(path)
    ws = wb.active
    return [row for row in ws.iter_rows(min_row=2, values_only=True)]


def write_xlsx(path: Path, headers: list, rows: list):
    """Записывает данные в Excel."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    print(f"Сохранено: {path}")


def append_xlsx(path: Path, headers: list, rows: list):
    """Создаёт или перезаписывает Excel — используется для промежуточного сохранения."""
    write_xlsx(path, headers, rows)