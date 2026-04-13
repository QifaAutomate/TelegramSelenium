import time
import random
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

INPUT_FILE = "files excel/leads.xlsx"
OUTPUT_FILE = "files excel/LEADS_FULL.xlsx"
CHAT_NAME = "WB Партнёры — чат"

options = Options()
options.add_argument(r"user-data-dir=C:\papka")

driver = webdriver.Chrome(options=options)
driver.get("https://web.telegram.org/a/")

wait = WebDriverWait(driver, 20)

def open_chat(name):
    try:
        search = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input#telegram-search-input")))
        search.click()
        search.send_keys(name)
        time.sleep(3)
        result = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".chat-item-clickable.search-result")))
        result.click()
        time.sleep(3)
        print(f"Чат '{name}' открыт")
        return True
    except Exception as e:
        print("Не удалось открыть чат:", e)
        return False

def get_messages():
    return driver.find_elements(
        By.CSS_SELECTOR,
        "div[id^='message-']:not([id^='message-group-'])"
    )

def get_first_message_id():
    messages = get_messages()
    return messages[0].get_attribute("id") if messages else None

def smart_scroll_up():
    try:
        chat = driver.find_element(By.CSS_SELECTOR, ".MessageList.custom-scroll")
        current = driver.execute_script("return arguments[0].scrollTop;", chat)
        for _ in range(5):
            current -= 300
            if current < 0:
                current = 0
            driver.execute_script("arguments[0].scrollTop = arguments[1];", chat, current)
            time.sleep(0.4)
        time.sleep(5)
    except Exception as e:
        print("Scroll error:", e)

def real_click(el):
    """Настоящий клик через ActionChains — не блокируется Telegram."""
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
    time.sleep(0.5)
    ActionChains(driver).move_to_element(el).click().perform()

def open_profile(sender_span):
    try:
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", sender_span)
        time.sleep(2)
        real_click(sender_span)
        print("  Клик по sender-title выполнен")

        time.sleep(4)

        opened = False
        for attempt in range(4):
            try:
                info_els = driver.find_elements(By.CSS_SELECTOR, ".MiddleHeader .info")
                if not info_els:
                    print(f"  Шапка не найдена (попытка {attempt + 1})")
                    time.sleep(2)
                    continue

                info_el = info_els[-1]
                real_click(info_el)
                print(f"  Клик по шапке чата (попытка {attempt + 1})")

                wait.until(EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#RightColumn .close-button")
                ))
                opened = True
                break
            except Exception as e:
                print(f"  Панель не открылась: {e}")
                time.sleep(2)

        if not opened:
            print("  Панель профиля не открылась")
            return False

        time.sleep(1.5)
        return True

    except Exception as e:
        print("open_profile error:", e)
        return False

def get_username():
    try:
        time.sleep(2)
        items = driver.find_elements(By.CSS_SELECTOR, "div.multiline-item span.title")
        for item in items:
            text = item.text.strip()
            if text.startswith("@"):
                return text
        return ""
    except:
        return ""

def get_PhoneNumber():
    try:
        time.sleep(2)
        items = driver.find_elements(By.CSS_SELECTOR, "div.multiline-item")
        for item in items:
            subtitle = item.find_element(By.CSS_SELECTOR, "span.subtitle")
            if subtitle.text.strip() == "Phone":
                title = item.find_element(By.CSS_SELECTOR, "span.title")
                return title.text.strip()
        return ""
    except:
        return ""

def close_profile_and_go_back():
    try:
        close_btn = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "#RightColumn .close-button")
        ))
        real_click(close_btn)
        print("  Профиль закрыт")
        time.sleep(2)
    except Exception as e:
        print("  Ошибка закрытия профиля:", e)

    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".chat-item-clickable")))
        chat_items = driver.find_elements(By.CSS_SELECTOR, ".chat-item-clickable")
        for item in chat_items:
            try:
                title_el = item.find_element(By.CSS_SELECTOR, "h3.fullName")
                if CHAT_NAME.lower() in title_el.text.lower():
                    real_click(item)
                    print("  Вернулись в групповой чат")
                    time.sleep(3)
                    break
            except:
                continue
    except Exception as e:
        print("  Ошибка возврата в чат:", e)

if not open_chat(CHAT_NAME):
    print("Не удалось открыть чат, завершаем")
    driver.quit()
    exit()

wb = load_workbook(INPUT_FILE)
ws = wb.active

sender_names = []
for row in ws.iter_rows(min_row=2, values_only=True):
    cell = row[0]
    if cell and cell not in sender_names:
        sender_names.append(cell)
    if len(sender_names) >= 50:
        break

results = {name: {"username": "", "phone": ""} for name in sender_names}
remaining = set(sender_names)

print(f"Уникальных отправителей: {len(sender_names)}")
print("Ищем:", remaining)

seen_ids = set()
no_growth_rounds = 0

for i in range(1000):
    before_id = get_first_message_id()
    messages = get_messages()

    print(f"\nРаунд {i}, сообщений: {len(messages)}, осталось: {len(remaining)}")

    for msg in messages:
        try:
            msg_id = msg.get_attribute("id")
            if not msg_id or msg_id in seen_ids:
                continue
            seen_ids.add(msg_id)

            sender_spans = msg.find_elements(By.CSS_SELECTOR, "span.sender-title")
            if not sender_spans:
                continue

            sender = sender_spans[0].text.strip()
            if not sender:
                continue

            matched = None
            for name in remaining:
                if name.lower() == sender.lower():
                    matched = name
                    break

            if not matched:
                continue

            print(f"  -> Найден: {matched} (msg_id: {msg_id})")

            if open_profile(sender_spans[0]):
                username = get_username()
                phone = get_PhoneNumber()
                print(f"  Username: {username if username else 'нет публичного юзернейма'}")
                print(f"  Phone: {phone if phone else 'нет номера телефона'}")
                results[matched]["username"] = username
                results[matched]["phone"] = phone
                remaining.discard(matched)

            close_profile_and_go_back()
            time.sleep(random.uniform(2, 4))
            no_growth_rounds = 0
            break

        except Exception as e:
            print("  msg error:", e)
            continue

    if not remaining:
        print("Все пользователи найдены!")
        break

    smart_scroll_up()

    after_id = get_first_message_id()
    if after_id == before_id:
        no_growth_rounds += 1
        print(f"Нет сдвига: {no_growth_rounds}/20")
    else:
        no_growth_rounds = 0

    if no_growth_rounds >= 20:
        print("Достигнут верх чата")
        break

out_wb = Workbook()
out_ws = out_wb.active
out_ws.append(["Отправитель", "Username", "PhoneNumber"])

for name in sender_names:
    out_ws.append([name, results[name]["username"], results[name]["phone"]])

out_wb.save(OUTPUT_FILE)
print(f"\nГотово! Результат: {OUTPUT_FILE}")

driver.quit()